from openpyxl import Workbook
from datetime import datetime
import lib.util as util

config = util.retrieveFromYAML("config/config.yml")

class ExportManager:

    def __init__(self) -> None:
        self.row, self.col = 2,2

    def exportToExcel(self, header: list, data: dict) -> None:
        
        wb = Workbook()
        ws = wb.active
   
        for index, val in enumerate(header):

            ws.cell(row=1, column=1+index, value=val)
            
        for key, val in data.items():
            
            self.col = 1
            #Key
            ws.cell(row=self.row, column=1, value=key)
            self.col += 1

            #Values
            for fieldVals in val:
                self.__digest(fieldVals,ws.cell)

            self.row += 1

        now = datetime.now()
        dt_string = now.strftime("%d%m%Y_%H%M%S")

        wb.save(f"{config["config"]["output"]}report_{dt_string}.xlsx")

    def __digest(self, data: any, func, horizontalWrite = True) -> None:

        if type(data) is list:

            if len(data) == 0:
                func(row=self.row, column=self.col, value="[]")
                return

            for index, value in enumerate(data):
                self.__digest(value,func,index == len(data)-1)

        else:

            func(row=self.row, column=self.col, value=data)

            if horizontalWrite:
                self.col += 1
            else:
                self.row += 1